import csv
import io
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_csv(data: List[Dict[str, Any]], columns: List[str]) -> str:
    """Gera arquivo CSV a partir dos dados"""
    output = io.StringIO()
    
    # Cabeçalhos traduzidos
    headers_pt = {
        'etiqueta_inventario': 'Número da Contagem',
        'inventario_cod_texto': 'Número da Etiqueta',
        'part_number_text': 'Material',
        'planta_text': 'Planta',
        'quantidade': 'Quantidade',
        'zona_invent_no_text': 'Zona',
        'created_date': 'Data Criação',
        'modified_date': 'Data Modificação',
        'created_by': 'Criado Por',
        'created_by_email': 'Email',
        'lote': 'Lote'
    }
    
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction='ignore')
    
    # Escrever cabeçalho traduzido
    header_row = {col: headers_pt.get(col, col) for col in columns}
    writer.writerow(header_row)
    
    for row in data:
        # Filtrar apenas as colunas especificadas
        filtered_row = {col: row.get(col, '') for col in columns}
        writer.writerow(filtered_row)
    
    return output.getvalue()


def generate_excel(data: List[Dict[str, Any]], columns: List[str]) -> bytes:
    """Gera arquivo Excel a partir dos dados"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contagens"
    
    # Cabeçalhos traduzidos
    headers_pt = {
        'etiqueta_inventario': 'Número da Contagem',
        'inventario_cod_texto': 'Número da Etiqueta',
        'part_number_text': 'Material',
        'planta_text': 'Planta',
        'quantidade': 'Quantidade',
        'zona_invent_no_text': 'Zona',
        'created_date': 'Data Criação',
        'modified_date': 'Data Modificação',
        'created_by': 'Criado Por',
        'created_by_email': 'Email',
        'lote': 'Lote'
    }
    
    # Estilo do cabeçalho
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Escrever cabeçalhos
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = headers_pt.get(col, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Escrever dados
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = row_data.get(col)
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()
